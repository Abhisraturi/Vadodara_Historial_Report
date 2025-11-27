import os
import threading
import subprocess
import time
import gc
import sys
from datetime import datetime, timedelta

from pylogix import PLC
from pycomm3 import SLCDriver
from openpyxl import load_workbook
from dotenv import load_dotenv
import pyodbc

# ===================== LOAD ENV =====================
load_dotenv()

# ===================== LOGGING / LOG ROTATION =====================
LOG_DIR = "logs"
LOOP_INTERVAL = int(os.getenv("LOOP_INTERVAL_SEC", "5"))

if not os.path.isdir(LOG_DIR):
    os.makedirs(LOG_DIR, exist_ok=True)


def get_log_file_path():
    """Return current day's log file path."""
    today_str = datetime.now().strftime("%Y%m%d")
    return os.path.join(LOG_DIR, f"plc_reader_{today_str}.log")


def purge_old_logs(days_keep: int = 7):
    """Keep only last `days_keep` days of log files."""
    cutoff = datetime.now() - timedelta(days=days_keep)
    try:
        for fname in os.listdir(LOG_DIR):
            if not fname.startswith("plc_reader_") or not fname.endswith(".log"):
                continue
            date_str = fname[len("plc_reader_"):-4]  # between prefix and .log
            try:
                fdate = datetime.strptime(date_str, "%Y%m%d")
            except ValueError:
                continue
            if fdate < cutoff:
                full_path = os.path.join(LOG_DIR, fname)
                os.remove(full_path)
    except Exception as e:
        # Log directory issues should never crash the main service
        print(f"[LOG_PURGE_ERROR] {e}")


def log(msg: str):
    ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    line = f"{ts} | {msg}"
    print(line)
    try:
        log_path = get_log_file_path()
        with open(log_path, "a", encoding="utf-8") as f:
            f.write(line + "\n")
    except Exception as e:
        # logging must never crash the loop
        print(f"[LOG_WRITE_ERROR] {e}")


# ===================== CONFIG / EXCEL COLS =====================
COL_PLC         = 1
COL_TAGNAME     = 2
COL_TAGINDEX    = 3
COL_TAGTYPE     = 4
COL_TAGDATATYPE = 5
COL_TAGVALUE    = 6
COL_STATUS      = 7
COL_TIMESTAMP   = 8


# ===================== HELPER: NORMALIZE PLC NAME =====================
def normalize_plc_name(name):
    if not name:
        return None
    name = str(name).strip()
    name = name.replace("::", "").replace(" ", "")
    if not name.startswith("["):
        name = f"[{name}]"
    return name


# ===================== SQL CONNECTION =====================
def get_sql_connection():
    SQL_DRIVER   = os.getenv("SQL_DRIVER")
    SQL_SERVER   = os.getenv("SQL_SERVER")
    SQL_DATABASE = os.getenv("SQL_DATABASE")
    SQL_AUTH     = os.getenv("SQL_AUTH", "windows")

    if SQL_AUTH.lower() == "windows":
        conn_str = (
            f"DRIVER={{{SQL_DRIVER}}};"
            f"SERVER={SQL_SERVER};"
            f"DATABASE={SQL_DATABASE};"
            "Trusted_Connection=yes;"
        )
    else:
        SQL_USERNAME = os.getenv("SQL_USERNAME")
        SQL_PASSWORD = os.getenv("SQL_PASSWORD")
        conn_str = (
            f"DRIVER={{{SQL_DRIVER}}};"
            f"SERVER={SQL_SERVER};"
            f"DATABASE={SQL_DATABASE};"
            f"UID={SQL_USERNAME};PWD={SQL_PASSWORD};"
        )

    return pyodbc.connect(conn_str)


# ===================== LOAD PLC CONFIG FROM ENV =====================
def load_plc_config_from_env():
    """
    Reads PLC_*_TYPE and PLC_*_IP from env.
    Supports names like:
        PLC_PLANT_1_TYPE
        PLC_PLANT_1_IP
        PLC_::[P1_NH3_PLC]_TYPE
        PLC_::[P1_NH3_PLC]_IP
    """
    plcs = {}

    for key, value in os.environ.items():
        key_u = key.upper()
        if key_u.startswith("PLC_") and key_u.endswith("_TYPE"):
            raw_name = key_u.replace("PLC_", "").replace("_TYPE", "")
            plc_type = value.strip().upper()

            normalized_name = normalize_plc_name(raw_name)

            ip_key = f"PLC_{raw_name}_IP"
            ip_val = os.getenv(ip_key)
            if not ip_val:
                log(f"[WARN] Missing IP for PLC {raw_name}")
                continue

            plcs[normalized_name] = {
                "type": plc_type,
                "ip": ip_val.strip(),
            }

    log("DEBUG: PLC MAP FROM ENV")
    for name, info in plcs.items():
        log(f"  {name} => {info}")

    return plcs


# ===================== LOAD EXCEL TAG MAP =====================
def load_excel_tags():
    EXCEL_FILE  = os.getenv("EXCEL_FILE_MAIN")
    EXCEL_SHEET = os.getenv("EXCEL_SHEET_MAIN")

    wb = load_workbook(EXCEL_FILE)
    ws = wb[EXCEL_SHEET]

    tag_map = {}
    row = 2

    while True:
        plc_raw = ws.cell(row=row, column=COL_PLC).value
        tag_cell = ws.cell(row=row, column=COL_TAGNAME).value

        if not tag_cell:
            break

        plc_name = normalize_plc_name(plc_raw)
        tag_name = str(tag_cell).strip()

        if plc_name not in tag_map:
            tag_map[plc_name] = []

        tag_map[plc_name].append({
            "row": row,
            "tag_name": tag_name,
            "tag_index": ws.cell(row=row, column=COL_TAGINDEX).value,
            "tag_type": ws.cell(row=row, column=COL_TAGTYPE).value,
            "tag_dtype": ws.cell(row=row, column=COL_TAGDATATYPE).value,
        })

        row += 1

    return wb, ws, tag_map


# ===================== PING CHECK =====================
def ping(ip):
    try:
        res = subprocess.run(
            ["ping", "-n", "1", "-w", "700", ip],
            capture_output=True,
            text=True
        )
        return "TTL" in res.stdout
    except Exception as e:
        log(f"[PING_ERROR] {ip} -> {e}")
        return False


# ===================== MICROLOGIX READER =====================
def read_micro_logix(ip, tag_list, retries=3):
    for attempt in range(retries):
        try:
            with SLCDriver(ip) as plc:
                addrs = [t["tag_name"] for t in tag_list]
                resp = plc.read(*addrs)
                resp_list = resp if isinstance(resp, list) else [resp]

                results = []
                for r, t in zip(resp_list, tag_list):
                    results.append({
                        "value": r.value,
                        "status": "OK" if not r.error else str(r.error),
                        "tag": t
                    })
                return results
        except Exception as e:
            log(f"[ERROR][MicroLogix {ip}] Attempt {attempt+1}/{retries} → {e}")
            time.sleep(0.2)

    # If all retries failed:
    return [{"value": None, "status": "NO RESPONSE", "tag": t} for t in tag_list]


# ===================== LOGIX READER (Compact / Micro800) =====================
def read_logix(ip, tag_list, is_micro800, retries=3):
    for attempt in range(retries):
        try:
            with PLC() as comm:
                comm.IPAddress = ip
                comm.Micro800 = is_micro800

                tags = [t["tag_name"] for t in tag_list]
                resp_list = comm.Read(tags)

                results = []
                for r, t in zip(resp_list, tag_list):
                    results.append({
                        "value": r.Value,
                        "status": r.Status,
                        "tag": t
                    })
                return results

        except Exception as e:
            log(f"[ERROR][Logix {ip}] Attempt {attempt+1}/{retries} → {e}")
            time.sleep(0.2)

    return [{"value": None, "status": "NO RESPONSE", "tag": t} for t in tag_list]


# ===================== THREAD WORKER =====================
def plc_worker(plc_name, plc_info, tag_list, result_bucket):
    ip = plc_info["ip"]
    plc_type = plc_info["type"]

    if not ping(ip):
        log(f"[OFFLINE] PLC {plc_name} @ {ip} not pinging. Skipping safely.")
        for t in tag_list:
            result_bucket.append({
                "plc": plc_name,
                "value": None,
                "status": "OFFLINE",
                "tag": t
            })
        return

    if plc_type == "MICROLOGIX":
        results = read_micro_logix(ip, tag_list)
    else:
        results = read_logix(ip, tag_list, is_micro800=(plc_type == "MICRO800"))

    for r in results:
        r["plc"] = plc_name
        result_bucket.append(r)


# ===================== MAIN LOOP WITH DAILY RESTART =====================
def main_loop():
    plcs = load_plc_config_from_env()
    wb, ws, tag_map = load_excel_tags()
    SQL_TABLE = os.getenv("SQL_TABLE")

    last_restart_date = None

    while True:
        # --------- Daily Restart Check (2 AM) ----------
        now_check = datetime.now()
        if (
            now_check.hour == 2
            and now_check.minute == 0
            and (last_restart_date is None or last_restart_date != now_check.date())
        ):
            log("Daily scheduled restart at 02:00 triggered.")
            # Flush logs & GC before restart
            gc.collect()
            last_restart_date = now_check.date()
            # re-exec this script in-place
            python = sys.executable
            os.execv(python, [python] + sys.argv)

        # --------- Purge Old Logs (keep last 7 days) ----------
        purge_old_logs(days_keep=7)

        # --------- Start New Scan ----------
        now = datetime.now()
        ts = now.strftime("%Y-%m-%d %H:%M:%S")
        log("========== NEW SCAN ==========")

        conn = get_sql_connection()
        cursor = conn.cursor()

        insert_sql = f"""
            INSERT INTO {SQL_TABLE} (
                ReadTime, PLC, TagIndex, TagName, TagType, TagDataType, TagValue, Status
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?)
        """

        all_results = []
        threads = []

        # ---- Spawn PLC Threads ----
        for plc_name, plc_info in plcs.items():
            if plc_name not in tag_map:
                log(f"[SKIP] No tags configured for {plc_name}")
                continue

            t = threading.Thread(
                target=plc_worker,
                args=(plc_name, plc_info, tag_map[plc_name], all_results),
                daemon=True,
            )
            t.start()
            threads.append(t)

        # ---- Wait for all PLCs to complete ----
        for t in threads:
            t.join()

        # ---- Sort by TagIndex ----
        all_results = sorted(all_results, key=lambda r: r["tag"]["tag_index"])

        # ---- Write SQL + Excel ----
        for r in all_results:
            plc_name = r["plc"]
            tinfo    = r["tag"]
            value    = r["value"]
            status   = r["status"]

            tag_name  = tinfo["tag_name"]
            tag_index = tinfo["tag_index"]
            tag_type  = tinfo["tag_type"]
            tag_dtype = tinfo["tag_dtype"]

            if isinstance(value, bool):
                excel_value = 1 if value else 0
                value_str = str(excel_value)
            else:
                excel_value = value
                value_str = None if value is None else str(value)

            row = tinfo["row"]
            ws.cell(row=row, column=COL_TAGVALUE).value   = excel_value
            ws.cell(row=row, column=COL_STATUS).value     = status
            ws.cell(row=row, column=COL_TIMESTAMP).value  = ts

            cursor.execute(
                insert_sql,
                now,
                plc_name,
                tag_index,
                tag_name,
                tag_type,
                tag_dtype,
                value_str,
                status
            )

            log(f"[SQL] {plc_name} | Index={tag_index} | {tag_name}={value_str} | {status}")

        conn.commit()
        cursor.close()
        conn.close()

        wb.save(os.getenv("EXCEL_FILE_MAIN"))

        log("========== SCAN COMPLETE ==========")

        # --------- Cleanup & Garbage Collection ----------
        del all_results
        del threads
        gc.collect()
        log("[GC] Garbage collection complete")

        time.sleep(LOOP_INTERVAL)


if __name__ == "__main__":
    main_loop()
