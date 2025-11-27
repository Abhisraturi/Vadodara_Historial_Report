import os
from datetime import datetime
from pylogix import PLC
from pycomm3 import SLCDriver
from openpyxl import load_workbook
from dotenv import load_dotenv

load_dotenv()

COL_PLC         = 1
COL_TAGNAME     = 2
COL_TAGINDEX    = 3
COL_TAGTYPE     = 4
COL_TAGDATATYPE = 5


def normalize_excel_plc_name(name):
    if not name:
        return None
    name = str(name).strip()
    name = name.replace("::", "").replace(" ", "")
    if not name.startswith("["):
        name = f"[{name}]"
    return name


def load_plc_config_from_env():
    plcs = {}

    for key, value in os.environ.items():
        key_up = key.upper()

        if key_up.startswith("PLC_") and key_up.endswith("_TYPE"):
            base = key_up.replace("PLC_", "").replace("_TYPE", "").strip()
            plc_type = value.strip().upper()

            # MICROLOGIX → keep full name
            if plc_type == "MICROLOGIX":
                excel_name = f"[{base}]"
                ip_key = f"PLC_{base}_IP"
                plc_ip = os.getenv(ip_key)

                if not plc_ip:
                    print(f"[WARNING] Missing IP for MICROLOGIX: {base}")
                    continue

                plcs[excel_name] = {"type": plc_type, "ip": plc_ip.strip()}

            # COMPACT / MICRO800 → remove _PLC
            else:
                clean = base.replace("_PLC", "")
                excel_name = f"[{clean}]"
                ip_key = f"PLC_{clean}_IP"
                plc_ip = os.getenv(ip_key)

                if not plc_ip:
                    print(f"[WARNING] Missing IP for PLC: {clean}")
                    continue

                plcs[excel_name] = {"type": plc_type, "ip": plc_ip.strip()}

    print("\nDEBUG PLC CONFIG:")
    for k, v in plcs.items():
        print(" ", k, "=>", v)

    return plcs


def load_excel_tags():
    EXCEL_FILE = os.getenv("EXCEL_FILE_MAIN", "Tagname.xlsx")
    EXCEL_SHEET = os.getenv("EXCEL_SHEET_MAIN", "Sheet1")

    wb = load_workbook(EXCEL_FILE)
    ws = wb[EXCEL_SHEET]

    tag_map = {}
    row = 2

    while True:
        plc_cell = ws.cell(row=row, column=COL_PLC).value
        tag_cell = ws.cell(row=row, column=COL_TAGNAME).value
        if not tag_cell:
            break

        plc_name = normalize_excel_plc_name(plc_cell)
        tag_name = str(tag_cell).strip()

        if plc_name not in tag_map:
            tag_map[plc_name] = []

        tag_map[plc_name].append({
            "row": row,
            "tag_name": tag_name,
            "tag_index": ws.cell(row=row, column=COL_TAGINDEX).value,
            "tag_type": ws.cell(row=row, column=COL_TAGTYPE).value,
            "tag_dtype": ws.cell(row=row, column=COL_TAGDATATYPE).value,
        })

        row += 1

    return tag_map


def read_micro_logix(plc_name, plc_ip, tag_list):
    results = []
    print(f"\n → MICROLOGIX READ: {plc_name} @ {plc_ip}")

    try:
        with SLCDriver(plc_ip) as plc:
            addrs = [t["tag_name"] for t in tag_list]
            resp = plc.read(*addrs)
            resp_list = resp if isinstance(resp, list) else [resp]

            for r, tinfo in zip(resp_list, tag_list):
                results.append({
                    "plc_name": plc_name,
                    "tag": tinfo,
                    "value": r.value,
                    "status": "OK" if not r.error else str(r.error)
                })
    except Exception as e:
        print(f"[ERROR] MICROLOGIX {plc_name}: {e}")

    return results


def read_logix(plc_name, plc_info, tag_list):
    results = []
    print(f"\n → LOGIX READ: {plc_name} @ {plc_info['ip']} (Type={plc_info['type']})")

    try:
        with PLC() as comm:
            comm.IPAddress = plc_info["ip"]
            comm.Micro800 = (plc_info["type"] == "MICRO800")

            tag_names = [t["tag_name"] for t in tag_list]
            resp_list = comm.Read(tag_names)

            for resp, tinfo in zip(resp_list, tag_list):
                results.append({
                    "plc_name": plc_name,
                    "tag": tinfo,
                    "value": resp.Value,
                    "status": resp.Status
                })

    except Exception as e:
        print(f"[ERROR] LOGIX {plc_name}: {e}")

    return results


def main():

    plcs = load_plc_config_from_env()
    excel_tags = load_excel_tags()

    all_results = []

    for plc_name, plc_info in plcs.items():

        if plc_name not in excel_tags:
            print(f"[SKIP] No tags found for {plc_name}")
            continue

        tag_list = excel_tags[plc_name]

        if plc_info["type"] == "MICROLOGIX":
            res = read_micro_logix(plc_name, plc_info["ip"], tag_list)
        else:
            res = read_logix(plc_name, plc_info, tag_list)

        all_results.extend(res)

    # ============================
    # SHOW FULL REPORT
    # ============================
    print("\n\n================= FULL PLC READ REPORT =================\n")

    total = len(all_results)
    ok = sum(1 for r in all_results if r["status"] == "OK" and r["value"] is not None)
    failed = total - ok

    for r in all_results:
        tag = r["tag"]
        print(
            f"[{r['plc_name']}] "
            f"Index={tag['tag_index']} | Tag={tag['tag_name']} | "
            f"Value={r['value']} | Status={r['status']}"
        )

    print("\n========================================================")
    print(f"Total Tags Read  : {total}")
    print(f"Successful       : {ok}")
    print(f"Failed/None      : {failed}")
    print("========================================================\n")

    if failed > 0:
        print(" FAILED TAGS:")
        for r in all_results:
            if r["value"] is None or r["status"] != "OK":
                tag = r["tag"]
                print(
                    f"  -> [{r['plc_name']}] {tag['tag_name']} "
                    f"(Index={tag['tag_index']}) Status={r['status']}"
                )


if __name__ == "__main__":
    main()
